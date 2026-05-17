"""
06_prepare_au_state_cafes.py
============================
Process ABS data into per-state cafe density for Vis 05.

Inputs (in ../data/):
  - data_cube_2.xlsx                       (ABS CABEE 8165DC02)
  - population_states_and_territories.xlsx (ABS 3101.0 Table 4)

Strategy:
  1. From CABEE Table 1 (June 2025): extract rows where ANZSIC code = 4511
     (Cafes and Restaurants). Read the "Total" count for each state.
  2. From Population Data1: find the "Persons" series for each state at
     the 2025-06-01 snapshot. (Both datasets use 30 June 2025.)
  3. Compute density = cafes per 10,000 residents (derived data).
  4. Output a clean CSV joinable to the AU state map by state_code.

Notes / caveats baked into the data:
  - ANZSIC class 4511 bundles "Cafes" with "Restaurants"; ABS doesn't isolate
    cafes only. This will be acknowledged in the chart caption.
  - "Other Territories" (3 cafes) is excluded — no matching state polygon.

Output:
  - cafes_by_state.csv

Run:
  python 06_prepare_au_state_cafes.py
"""

from pathlib import Path
from datetime import datetime
import openpyxl
import csv

SCRIPT_DIR = Path(__file__).parent.resolve()
DATA_DIR = SCRIPT_DIR.parent / "data"

CABEE_FILE = DATA_DIR / "data_cube_2.xlsx"
POP_FILE = DATA_DIR / "population_states_and_territories.xlsx"

# Map ABS full state names → standard 2-3 letter codes
STATE_CODE = {
    "New South Wales": "NSW",
    "Victoria": "VIC",
    "Queensland": "QLD",
    "South Australia": "SA",
    "Western Australia": "WA",
    "Tasmania": "TAS",
    "Northern Territory": "NT",
    "Australian Capital Territory": "ACT",
}

# Population columns in Data1 sheet (Persons × state, 0-indexed)
POP_COLS = {
    "NSW": 19, "VIC": 20, "QLD": 21, "SA": 22,
    "WA": 23, "TAS": 24, "NT": 25, "ACT": 26,
}

SNAPSHOT_DATE = datetime(2025, 6, 1)  # 30 June 2025 ≈ 2025-06-01 row


def load_cafe_counts() -> dict:
    """Extract Cafes & Restaurants (ANZSIC 4511) total counts per state."""
    print(f"[1/3] Reading {CABEE_FILE.name} …")
    wb = openpyxl.load_workbook(CABEE_FILE, read_only=True, data_only=True)
    ws = wb["Table 1"]

    counts = {}
    for row in ws.iter_rows(min_row=8, values_only=True):
        if row[1] == 4511:  # ANZSIC class column
            state_full = row[0]
            total = row[8]  # "Total" column (Operating at start of FY)
            if state_full in STATE_CODE and total is not None:
                code = STATE_CODE[state_full]
                counts[code] = int(total)

    print(f"      ✓ Found cafes for {len(counts)} states")
    return counts


def load_populations() -> dict:
    """Extract Persons population per state at June 2025."""
    print(f"[2/3] Reading {POP_FILE.name} …")
    wb = openpyxl.load_workbook(POP_FILE, read_only=True, data_only=True)
    ws = wb["Data1"]

    pops = {}
    for row in ws.iter_rows(min_row=11, values_only=True):
        date_val = row[0]
        if isinstance(date_val, datetime) and date_val == SNAPSHOT_DATE:
            for code, col in POP_COLS.items():
                pops[code] = int(row[col])
            break

    print(f"      ✓ Found populations for {len(pops)} states at {SNAPSHOT_DATE.date()}")
    return pops


def main():
    cafes = load_cafe_counts()
    pops = load_populations()

    # Sanity check — both datasets should have the same 8 states
    missing = set(STATE_CODE.values()) - set(cafes.keys()) - set(pops.keys())
    if missing:
        print(f"⚠️  Missing data for: {missing}")

    print("[3/3] Computing per-capita density …")
    rows = []
    for state_full, code in STATE_CODE.items():
        if code in cafes and code in pops:
            density = cafes[code] / pops[code] * 10000
            rows.append({
                "state": state_full,
                "state_code": code,
                "cafes_restaurants": cafes[code],
                "population": pops[code],
                "density_per_10k": round(density, 2),
            })

    # Sort by density desc (so the CSV reads naturally)
    rows.sort(key=lambda r: -r["density_per_10k"])

    out = DATA_DIR / "cafes_by_state.csv"
    with out.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)

    print(f"\n✓ Saved {out.name} ({out.stat().st_size} bytes, {len(rows)} rows)")

    # Preview
    print("\n--- Preview (sorted by density) ---")
    print(f"{'State':5s}  {'Cafes':>7s}  {'Pop':>11s}  {'Per 10k':>8s}")
    for r in rows:
        print(
            f"{r['state_code']:5s}  {r['cafes_restaurants']:>7,}  "
            f"{r['population']:>11,}  {r['density_per_10k']:>8.2f}"
        )


if __name__ == "__main__":
    main()
